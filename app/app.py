
from fastapi import FastAPI, Form, Request
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from io import BytesIO
import datetime

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

@app.get("/")
def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/generate")
async def generate_excel(
    request: Request,
    datum: str = Form(...),
    bauort: str = Form(...),
    bf: str = Form(...),
    beschreibung: str = Form(...),
    geraet: str = Form(""),
    nachname1: str = Form(...),
    vorname1: str = Form(...),
    ausweis1: str = Form(...),
    beginn1: str = Form(...),
    ende1: str = Form(...),
    nachname2: str = Form(""),
    vorname2: str = Form(""),
    ausweis2: str = Form(""),
    beginn2: str = Form(""),
    ende2: str = Form(""),
    nachname3: str = Form(""),
    vorname3: str = Form(""),
    ausweis3: str = Form(""),
    beginn3: str = Form(""),
    ende3: str = Form(""),
    nachname4: str = Form(""),
    vorname4: str = Form(""),
    ausweis4: str = Form(""),
    beginn4: str = Form(""),
    ende4: str = Form(""),
    nachname5: str = Form(""),
    vorname5: str = Form(""),
    ausweis5: str = Form(""),
    beginn5: str = Form(""),
    ende5: str = Form(""),
):
    wb = load_workbook("GP-t.xlsx")
    ws = wb.active

    ws.cell(row=4, column=4).value = datum  # D4 helyett, hogy ne legyen MergedCell error
    ws["D5"] = bauort
    ws["D6"] = bf
    ws["D7"] = beschreibung
    ws["D8"] = geraet

    # Itt jönnek a dolgozók, de nem változtattam rajtuk most

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Arbeitsnachweis_{datetime.date.today().isoformat()}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
