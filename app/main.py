# main.py
from fastapi import FastAPI, Request, Form, Response, Body
from fastapi.responses import HTMLResponse, PlainTextResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from datetime import datetime, time
from io import BytesIO
import os
import uuid
import textwrap
import traceback
import json
import sqlite3
from pathlib import Path
from typing import Tuple, Optional, List

# Hálózat/HTTP
import httpx
import socket
import re
from urllib.parse import urlparse, urlencode

# ---- Google Drive (service account) ----
GDRIVE_JSON = os.getenv("GDRIVE_SERVICE_ACCOUNT_JSON", "")
GDRIVE_FOLDER_ID = os.getenv("GDRIVE_FOLDER_ID", "").strip()

DRIVE_ENABLED = bool(GDRIVE_JSON and GDRIVE_FOLDER_ID)
drive_svc = None
if DRIVE_ENABLED:
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
        creds = service_account.Credentials.from_service_account_info(
            json.loads(GDRIVE_JSON),
            scopes=["https://www.googleapis.com/auth/drive.file"],
        )
        drive_svc = build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception as e:
        print("Drive init failed:", repr(e))
        DRIVE_ENABLED = False
        drive_svc = None

def drive_upload_bytes(filename: str, data: bytes, mime: str) -> Optional[str]:
    if not (DRIVE_ENABLED and drive_svc):
        return None
    try:
        media = MediaIoBaseUpload(BytesIO(data), mimetype=mime, resumable=False)
        meta = {"name": filename, "parents": [GDRIVE_FOLDER_ID]}
        file = drive_svc.files().create(body=meta, media_body=media, fields="id").execute()
        return file.get("id")
    except Exception as e:
        print("Drive upload failed:", repr(e))
        return None

# === Drive helper: DB tükrözés ===
DB_DRIVE_NAME = "app.db"
DB_DRIVE_ID: Optional[str] = None

def drive_find_file_id_by_name(name: str) -> Optional[str]:
    """Megkeresi a fájlt a mappában név alapján, visszaadja az id-t vagy None."""
    if not (DRIVE_ENABLED and drive_svc):
        return None
    try:
        q = f"name = '{name}' and '{GDRIVE_FOLDER_ID}' in parents and trashed = false"
        res = drive_svc.files().list(q=q, fields="files(id,name)", pageSize=1).execute()
        files = res.get("files", [])
        if files:
            return files[0]["id"]
        return None
    except Exception as e:
        print("Drive find file failed:", repr(e))
        return None

def drive_download_file(file_id: str) -> Optional[bytes]:
    """Letölti a fájlt byte-okba."""
    if not (DRIVE_ENABLED and drive_svc and file_id):
        return None
    try:
        req = drive_svc.files().get_media(fileId=file_id)
        buf = BytesIO()
        downloader = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        return buf.getvalue()
    except Exception as e:
        print("Drive download failed:", repr(e))
        return None

def drive_upload_or_update(filename: str, data: bytes, mime: str, existing_id: Optional[str]) -> Optional[str]:
    """Ha van existing_id -> update, különben create."""
    if not (DRIVE_ENABLED and drive_svc):
        return None
    try:
        media = MediaIoBaseUpload(BytesIO(data), mimetype=mime, resumable=False)
        if existing_id:
            file = drive_svc.files().update(fileId=existing_id, media_body=media).execute()
            return file.get("id")
        else:
            meta = {"name": filename, "parents": [GDRIVE_FOLDER_ID]}
            file = drive_svc.files().create(body=meta, media_body=media, fields="id").execute()
            return file.get("id")
    except Exception as e:
        print("Drive upload/update failed:", repr(e))
        return None

# ---- PDF (ReportLab) - opcionális előnézet ----
REPORTLAB_AVAILABLE = False
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# ---- Képgenerálás (Pillow) ----
try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

# ---- App / statikus / sablonok ----
app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# ---- Session a bejelentkezéshez (NYELVHEZ NEM HASZNÁLJUK) ----
SESSION_SECRET = os.getenv("SESSION_SECRET", "change-me-dev-secret")
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET, same_site="lax")

# ---- Login beállítások ----
APP_USERNAME = os.getenv("APP_USERNAME", "user")
APP_PASSWORD = os.getenv("APP_PASSWORD", "user")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin")

def _is_user(request: Request) -> bool:
    return bool(request.session.get("auth_ok") is True)

def _is_admin(request: Request) -> bool:
    return bool(request.session.get("admin_ok") is True)

# ---- Tárolók / DB init ----
BASE_DIR = Path(os.getcwd())
DATA_DIR = BASE_DIR / "data"
GEN_DIR = BASE_DIR / "generated"
DATA_DIR.mkdir(exist_ok=True)
GEN_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "app.db"

# >>> Induláskor megpróbáljuk a DB-t letölteni a Drive-ról
def try_sync_db_from_drive():
    global DB_DRIVE_ID
    if not DRIVE_ENABLED:
        print("Drive DB sync: disabled")
        return
    try:
        fid = drive_find_file_id_by_name(DB_DRIVE_NAME)
        if not fid:
            print("Drive DB sync: no remote app.db yet")
            return
        data = drive_download_file(fid)
        if data:
            DB_PATH.write_bytes(data)
            DB_DRIVE_ID = fid
            print(f"Drive DB sync: downloaded {DB_DRIVE_NAME} ({len(data)} bytes)")
        else:
            print("Drive DB sync: download returned no data")
    except Exception as e:
        print("Drive DB sync (download) failed:", repr(e))

try_sync_db_from_drive()

def db_conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    with db_conn() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            datum TEXT,
            bau TEXT,
            basf_beauftragter TEXT,
            beschreibung TEXT,
            break_minutes INTEGER,
            vorname1 TEXT, nachname1 TEXT, ausweis1 TEXT, beginn1 TEXT, ende1 TEXT, vorhaltung1 TEXT,
            vorname2 TEXT, nachname2 TEXT, ausweis2 TEXT, beginn2 TEXT, ende2 TEXT, vorhaltung2 TEXT,
            vorname3 TEXT, nachname3 TEXT, ausweis3 TEXT, beginn3 TEXT, ende3 TEXT, vorhaltung3 TEXT,
            vorname4 TEXT, nachname4 TEXT, ausweis4 TEXT, beginn4 TEXT, ende4 TEXT, vorhaltung4 TEXT,
            vorname5 TEXT, nachname5 TEXT, ausweis5 TEXT, beginn5 TEXT, ende5 TEXT, vorhaltung5 TEXT,
            excel_filename TEXT,
            payload_json TEXT
        )
        """)
init_db()

def sync_db_to_drive():
    """Minden írás után töltsük fel/frissítsük a DB-t a Drive mappába."""
    global DB_DRIVE_ID
    if not DRIVE_ENABLED:
        return
    try:
        data = DB_PATH.read_bytes()
        new_id = drive_upload_or_update(DB_DRIVE_NAME, data, "application/octet-stream", DB_DRIVE_ID)
        if new_id:
            DB_DRIVE_ID = new_id
            print(f"Drive DB sync: uploaded {DB_DRIVE_NAME} ({len(data)} bytes), id={DB_DRIVE_ID}")
        else:
            print("Drive DB sync: upload returned no id")
    except Exception as e:
        print("Drive DB sync (upload) failed:", repr(e))

# ---------- helpers (Excel stb.) ----------
def merged_ranges(ws):
    return [(r.min_row, r.min_col, r.max_row, r.max_col) for r in ws.merged_cells.ranges]

def in_range(rng, r, c):
    r1, c1, r2, c2 = rng
    return (r1 <= r <= r2) and (c1 <= c <= c2)

def block_of(ws, r, c):
    for rng in merged_ranges(ws):
        if in_range(rng, r, c):
            return rng
    return (r, c, r, c)

def top_left_of_block(ws, r, c):
    r1, c1, _, _ = block_of(ws, r, c)
    return r1, c1

def set_text(ws, r, c, text, wrap=False, align_left=False, valign_top=False):
    rr, cc = top_left_of_block(ws, r, c)
    cell = ws.cell(row=rr, column=cc)
    cell.value = text
    cell.alignment = Alignment(
        wrap_text=wrap,
        horizontal=("left" if align_left else "center"),
        vertical=("top" if valign_top else cell.alignment.vertical or "center"),
    )

def set_text_addr(ws, addr, text, *, wrap=False, horizontal="left", vertical="center"):
    cell = ws[addr]
    rr, cc = top_left_of_block(ws, cell.row, cell.column)
    tgt = ws.cell(row=rr, column=cc)
    tgt.value = text
    tgt.alignment = Alignment(wrap_text=wrap, horizontal=horizontal, vertical=vertical)

def parse_hhmm(s: str) -> Optional[time]:
    s = (s or "").strip()
    if not s:
        return None
    hh, mm = s.split(":")
    return time(int(hh), int(mm))

def overlap_minutes(a1: time, a2: time, b1: time, b2: time) -> int:
    dt = datetime(2000,1,1)
    A1 = dt.replace(hour=a1.hour, minute=a1.minute)
    A2 = dt.replace(hour=a2.hour, minute=a2.minute)
    B1 = dt.replace(hour=b1.hour, minute=b1.minute)
    B2 = dt.replace(hour=b2.hour, minute=b2.minute)
    start = max(A1, B1)
    end   = min(A2, B2)
    if end <= start:
        return 0
    return int((end - start).total_seconds() // 60)

def hours_with_breaks(beg: Optional[time], end: Optional[time], pause_min: int = 60) -> float:
    if not beg or not end:
        return 0.0
    dt = datetime(2000,1,1)
    start = dt.replace(hour=beg.hour, minute=beg.minute)
    finish = dt.replace(hour=end.hour, minute=end.minute)
    if finish <= start:
        return 0.0
    total_min = int((finish - start).total_seconds() // 60)
    if pause_min >= 60:
        minus = overlap_minutes(beg, end, time(9,0), time(9,15)) \
              + overlap_minutes(beg, end, time(12,0), time(12,45))
    else:
        minus = min(total_min, 30)
    return max(0.0, (total_min - minus) / 60.0)

def find_header_positions(ws):
    pos = {}
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=120):
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                t = v.strip()
                if t == "Name":
                    pos["name_col"] = cell.column; header_row = cell.row
                if t == "Vorname":
                    pos["vorname_col"] = cell.column
                if "Ausweis" in t or "Kennzeichen" in t:
                    pos["ausweis_col"] = cell.column
                if t == "Beginn":
                    pos["beginn_col"] = cell.column; pos["subheader_row"] = cell.row
                if t == "Ende":
                    pos["ende_col"] = cell.column
                if "Anzahl Stunden" in t:
                    pos["stunden_col"] = cell.column
                if t.lower().startswith("vorhaltung") or "beauftragtes gerät" in t.lower():
                    pos["vorhaltung_col"] = cell.column
        if all(k in pos for k in ["name_col","vorname_col","ausweis_col","beginn_col","ende_col","stunden_col","subheader_row"]):
            break
    pos["data_start_row"] = pos.get("subheader_row", header_row) + 1
    return pos

def find_total_cells(ws, stunden_col):
    total_row = None
    right_of_label = None
    for row in ws.iter_rows(min_row=1, max_row=200):
        for cell in row:
            if isinstance(cell.value, str) and "Gesamtstunden" in cell.value:
                total_row = cell.row
                r_neighbor, c_neighbor = total_row, cell.column + 1
                rr, cc = top_left_of_block(ws, r_neighbor, c_neighbor)
                right_of_label = (rr, cc)
                break
        if total_row:
            break
    stunden_total = None
    if total_row:
        rr, cc = top_left_of_block(ws, total_row, stunden_col)
        stunden_total = (rr, cc)
    return right_of_label, stunden_total

def find_description_block(ws) -> Tuple[int,int,int,int]:
    return (6, 1, 15, 7)  # A6..G15

def _excel_col_width_to_pixels(width):
    if width is None:
        width = 8.43
    return int(round(7 * width + 5))

def _excel_row_height_to_pixels(height):
    if height is None:
        height = 15.0
    return int(round(height * 96 / 72))

def _get_block_pixel_size(ws, r1, c1, r2, c2):
    w_px = 0
    for c in range(c1, c2 + 1):
        letter = get_column_letter(c)
        cd = ws.column_dimensions.get(letter)
        w_px += _excel_col_width_to_pixels(getattr(cd, "width", None))
    h_px = 0
    for r in range(r1, r2 + 1):
        rd = ws.row_dimensions.get(r)
        h_px += _excel_row_height_to_pixels(getattr(rd, "height", None))
    w_px = max(40, w_px)
    h_px = max(40, h_px)
    return w_px, h_px

def _get_col_pixel_width(ws, col_index):
    letter = get_column_letter(col_index)
    cd = ws.column_dimensions.get(letter)
    return _excel_col_width_to_pixels(getattr(cd, "width", None))

LEFT_INSET_PX = 25
BOTTOM_CROP   = 0.92

def _make_description_image(text, w_px, h_px):
    if not PIL_AVAILABLE:
        raise RuntimeError("PIL not available")
    img = PILImage.new("RGB", (w_px, h_px), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    font = None
    for name in ["arial.ttf", "Arial.ttf", "DejaVuSans.ttf", "LiberationSans-Regular.ttf"]:
        try:
            font = ImageFont.truetype(name, 14); break
        except Exception:
            font = None
    if font is None:
        font = ImageFont.load_default()

    pad_left, pad_top, pad_right, pad_bottom = 12, 10, 0, 10
    avail_w = max(10, w_px - (pad_left + pad_right))
    avail_h = max(10, h_px - (pad_top + pad_bottom))

    sample = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 .,:;-"
    avg_w = max(6, sum(draw.textlength(ch, font=font) for ch in sample) / len(sample))
    avg_w_eff = avg_w * 0.90
    max_chars_per_line = max(10, int(avail_w / avg_w_eff))

    paragraphs = (text or "").replace("\r\n","\n").replace("\r","\n").split("\n")
    lines = []
    for para in paragraphs:
        if not para:
            lines.append(""); continue
        lines.extend(textwrap.wrap(para, width=max_chars_per_line, replace_whitespace=False, break_long_words=False))

    ascent, descent = font.getmetrics()
    line_h = ascent + descent + 4
    max_lines = max(1, int(avail_h // line_h))
    if len(lines) > max_lines:
        lines = lines[:max_lines - 1] + ["…"]

    x = pad_left; y = pad_top
    for ln in lines:
        draw.text((x, y), ln, fill=(0, 0, 0), font=font)
        y += line_h
    return img

def _xlimage_from_pil(pil_img):
    buf = BytesIO(); pil_img.save(buf, format="PNG"); buf.seek(0); return XLImage(buf)

def insert_description_as_image(ws, r1, c1, r2, c2, text):
    if not PIL_AVAILABLE:
        print("IMG: PIL not available"); return False
    try:
        block_w_px, block_h_px = _get_block_pixel_size(ws, r1, c1, r2, c2)
        colA_w_px = _get_col_pixel_width(ws, 1)
        anchor_col = c1 + 1
        anchor = f"{get_column_letter(anchor_col)}{r1}"
        new_w_px = max(40, block_w_px - colA_w_px - LEFT_INSET_PX)
        new_h_px = int(block_h_px * BOTTOM_CROP)
        pil_img = _make_description_image(text or "", new_w_px, new_h_px)
        xlimg = _xlimage_from_pil(pil_img)
        ws.add_image(xlimg, anchor)
        set_text(ws, r1, c1, "", wrap=False, align_left=True, valign_top=True)
        return True
    except Exception as e:
        print("IMG insert failed:", repr(e)); traceback.print_exc(); return False

def set_print_defaults(ws):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 0
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = 95
    if hasattr(ws, "sheet_properties") and hasattr(ws.sheet_properties, "pageSetUpPr"):
        ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)
    last_data_row = 1; last_data_col = 1
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                last_data_row = max(last_data_row, r); last_data_col = max(last_data_col, c)
    last_col_letter = get_column_letter(last_data_col)
    ws.print_area = f"A1:{last_col_letter}{last_data_row}"
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False

def _build_pdf_preview(date_text, bau, basf_beauftragter, beschreibung, ws, r1, c1, r2, c2, workers, total_hours):
    if not (REPORTLAB_AVAILABLE and PIL_AVAILABLE):
        raise RuntimeError("ReportLab/PIL missing")
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader

    pw, ph = landscape(A4)
    margin_left = 12 * mm; margin_right = 12 * mm; margin_top = 12 * mm; margin_bottom = 12 * mm
    buf = BytesIO(); c = canvas.Canvas(buf, pagesize=(pw, ph))
    y = ph - margin_top

    c.setFont("Helvetica-Bold", 12); c.drawString(margin_left, y, f"Datum: {date_text}"); y -= 16
    c.setFont("Helvetica", 12); c.drawString(margin_left, y, f"Bau: {bau}"); y -= 16
    if (basf_beauftragter or "").strip():
        c.drawString(margin_left, y, f"BASF Beauftragter: {basf_beauftragter}"); y -= 10
    y -= 6

    block_w_px, block_h_px = _get_block_pixel_size(ws, r1, c1, r2, c2)
    colA_w_px = _get_col_pixel_width(ws, 1)
    new_w_px = max(40, block_w_px - colA_w_px - LEFT_INSET_PX)
    new_h_px = int(block_h_px * BOTTOM_CROP)
    pil_img = _make_description_image(beschreibung or "", new_w_px, new_h_px)
    w_pt = new_w_px * 0.75; h_pt = new_h_px * 0.75
    max_w_pt = pw - margin_left - margin_right
    if w_pt > max_w_pt:
        scale = max_w_pt / w_pt; w_pt *= scale; h_pt *= scale
    y -= h_pt
    c.drawImage(ImageReader(pil_img), margin_left, y, width=w_pt, height=h_pt, preserveAspectRatio=True, mask='auto'); y -= 12

    c.setFont("Helvetica-Bold", 11)
    headers = ["Name", "Vorname", "Ausweis", "Beginn", "Ende", "Stunden", "Vorhaltung"]
    col_widths = [70, 70, 90, 60, 60, 60, 100]
    x = margin_left
    for hdr, w in zip(headers, col_widths):
        c.drawString(x, y, hdr); x += w
    y -= 14; c.setLineWidth(0.3); c.line(margin_left, y, margin_left + sum(col_widths), y); y -= 6

    c.setFont("Helvetica", 10)
    for (vn, nn, aw, bg, en, vh) in workers:
        x = margin_left
        vals = [nn, vn, aw, bg, en, f"{hours_with_breaks(parse_hhmm(bg), parse_hhmm(en)):.2f}", vh]
        for val, w in zip(vals, col_widths):
            c.drawString(x, y, str(val or "")); x += w
        y -= 14
        if y < margin_bottom + 40:
            c.showPage(); y = ph - margin_top; c.setFont("Helvetica", 10)

    y -= 6; c.setFont("Helvetica-Bold", 11); c.drawString(margin_left, y, f"Gesamtstunden: {total_hours:.2f}")
    c.showPage(); c.save(); buf.seek(0); return buf.read()

# ---------- User Login / Logout ----------
@app.get("/login", response_class=HTMLResponse)
async def login_form(request: Request, next: str = "/"):
    if _is_user(request):
        return RedirectResponse(next or "/", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "error": False, "next": next})

@app.post("/login", response_class=HTMLResponse)
async def login_submit(request: Request, username: str = Form(...), password: str = Form(...), next: str = Form("/")):
    if username == APP_USERNAME and password == APP_PASSWORD:
        request.session.clear(); request.session["auth_ok"] = True
        return RedirectResponse(next or "/", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "error": True, "next": next}, status_code=401)

@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)

# ---------- Admin Login / Logout ----------
@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_form(request: Request, next: str = "/admin"):
    if _is_admin(request):
        return RedirectResponse(next or "/admin", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "error": False, "next": next})

@app.post("/admin/login", response_class=HTMLResponse)
async def admin_login_submit(request: Request, username: str = Form(...), password: str = Form(...), next: str = Form("/admin")):
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        request.session.clear(); request.session["admin_ok"] = True
        return RedirectResponse(next or "/admin", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "error": True, "next": next}, status_code=401)

@app.get("/admin/logout")
async def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse("/admin/login", status_code=303)

# ---------- Főoldal (NYELV: csak queryből; default: de) ----------
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    if not _is_user(request):
        return RedirectResponse("/login?next=/", status_code=303)
    lang = (request.query_params.get("lang") or "").strip().lower()
    if lang not in {"de", "hr"}:
        lang = "de"
    return templates.TemplateResponse("index.html", {"request": request, "lang": lang})

# ---------- Excel generálás + DB mentés + Drive feltöltés ----------
@app.post("/generate_excel")
async def generate_excel(
    request: Request,
    datum: str = Form(...),
    bau: str = Form(...),
    basf_beauftragter: str = Form(""),
    geraet: str = Form(""),
    beschreibung: str = Form(""),
    break_minutes: int = Form(60),
    vorname1: str = Form(""), nachname1: str = Form(""), ausweis1: str = Form(""), beginn1: str = Form(""), ende1: str = Form(""), vorhaltung1: str = Form(""),
    vorname2: str = Form(""), nachname2: str = Form(""), ausweis2: str = Form(""), beginn2: str = Form(""), ende2: str = Form(""), vorhaltung2: str = Form(""),
    vorname3: str = Form(""), nachname3: str = Form(""), ausweis3: str = Form(""), beginn3: str = Form(""), ende3: str = Form(""), vorhaltung3: str = Form(""),
    vorname4: str = Form(""), nachname4: str = Form(""), ausweis4: str = Form(""), beginn4: str = Form(""), ende4: str = Form(""), vorhaltung4: str = Form(""),
    vorname5: str = Form(""), nachname5: str = Form(""), ausweis5: str = Form(""), beginn5: str = Form(""), ende5: str = Form(""), vorhaltung5: str = Form(""),
):
    if not _is_user(request):
        return RedirectResponse("/login?next=/", status_code=303)

    wb = load_workbook(os.path.join(os.getcwd(), "GP-t.xlsx")); ws = wb.active

    date_text = datum
    try:
        dt = datetime.strptime(datum.strip(), "%Y-%m-%d"); date_text = dt.strftime("%d.%m.%Y")
    except Exception:
        pass
    set_text_addr(ws, "B2", date_text, horizontal="left")
    set_text_addr(ws, "B3", bau,        horizontal="left")
    if (basf_beauftragter or "").strip():
        set_text_addr(ws, "E3", basf_beauftragter, horizontal="left")

    r1, c1, r2, c2 = find_description_block(ws)
    for r in range(r1, r2 + 1):
        if ws.row_dimensions.get(r) is None or ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 22

    inserted = False
    text_in = (beschreibung or "").strip()
    if text_in:
        inserted = insert_description_as_image(ws, r1, c1, r2, c2, text_in)
    if not inserted:
        set_text(ws, r1, c1+1, text_in, wrap=True, align_left=True, valign_top=True)

    pos = find_header_positions(ws); row = pos["data_start_row"]
    vorhaltung_col = pos.get("vorhaltung_col", None)

    workers: List[Tuple[str,str,str,str,str,str]] = []
    for i in range(1, 6):
        vn = locals().get(f"vorname{i}", "") or ""
        nn = locals().get(f"nachname{i}", "") or ""
        aw = locals().get(f"ausweis{i}", "") or ""
        bg = locals().get(f"beginn{i}", "") or ""
        en = locals().get(f"ende{i}", "") or ""
        vh = locals().get(f"vorhaltung{i}", "") or ""
        if not (vn or nn or aw or bg or en or vh): continue
        workers.append((vn, nn, aw, bg, en, vh))

    total_hours = 0.0
    for (vn, nn, aw, bg, en, vh) in workers:
        set_text(ws, row, pos["name_col"], nn, wrap=False, align_left=True)
        set_text(ws, row, pos["vorname_col"], vn, wrap=False, align_left=True)
        set_text(ws, row, pos["ausweis_col"], aw, wrap=False, align_left=True)
        set_text(ws, row, pos["beginn_col"], bg, wrap=False, align_left=True)
        set_text(ws, row, pos["ende_col"], en, wrap=False, align_left=True)
        if vorhaltung_col and (vh or "").strip():
            set_text(ws, row, vorhaltung_col, vh, wrap=True, align_left=True, valign_top=True)
        hb = parse_hhmm(bg); he = parse_hhmm(en)
        h = round(hours_with_breaks(hb, he, int(break_minutes)), 2)
        total_hours += h; set_text(ws, row, pos["stunden_col"], h, wrap=False, align_left=True); row += 1

    right_of_label, stunden_total = find_total_cells(ws, pos["stunden_col"])
    if stunden_total:
        tr, tc = stunden_total; set_text(ws, tr, tc, round(total_hours, 2), wrap=False, align_left=True)
    if right_of_label:
        rr, rc = right_of_label; set_text(ws, rr, rc, "", wrap=False, align_left=True)

    try:
        set_print_defaults(ws)
    except Exception as e:
        print("PRINT SETUP WARN:", repr(e))

    excel_name = f"leistungsnachweis_{uuid.uuid4().hex[:8]}.xlsx"
    excel_path = GEN_DIR / excel_name
    wb.save(excel_path.as_posix())
    with open(excel_path, "rb") as f:
        excel_bytes = f.read()

    drive_file_id = drive_upload_bytes(
        filename=excel_name,
        data=excel_bytes,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    payload = {
        "datum": datum, "bau": bau, "basf_beauftragter": basf_beauftragter,
        "beschreibung": beschreibung, "break_minutes": int(break_minutes),
        "workers": [
            {"vorname": locals().get(f"vorname{i}", ""), "nachname": locals().get(f"nachname{i}", ""),
             "ausweis": locals().get(f"ausweis{i}", ""), "beginn": locals().get(f"beginn{i}", ""),
             "ende": locals().get(f"ende{i}", ""), "vorhaltung": locals().get(f"vorhaltung{i}", "")}
            for i in range(1, 6)
        ],
        "drive_file_id": drive_file_id
    }
    with db_conn() as c:
        c.execute("""
            INSERT INTO submissions (
                created_at, datum, bau, basf_beauftragter, beschreibung, break_minutes,
                vorname1, nachname1, ausweis1, beginn1, ende1, vorhaltung1,
                vorname2, nachname2, ausweis2, beginn2, ende2, vorhaltung2,
                vorname3, nachname3, ausweis3, beginn3, ende3, vorhaltung3,
                vorname4, nachname4, ausweis4, beginn4, ende4, vorhaltung4,
                vorname5, nachname5, ausweis5, beginn5, ende5, vorhaltung5,
                excel_filename, payload_json
            ) VALUES (
                ?,?,?,?,?,?,
                ?,?,?,?,?,?,
                ?,?,?,?,?,?,
                ?,?,?,?,?,?,
                ?,?,?,?,?,?,
                ?,?,?,?,?,?,
                ?,?
            )
        """, (
            datetime.utcnow().isoformat(),
            datum, bau, basf_beauftragter, beschreibung, int(break_minutes),
            locals().get("vorname1",""), locals().get("nachname1",""), locals().get("ausweis1",""), locals().get("beginn1",""), locals().get("ende1",""), locals().get("vorhaltung1",""),
            locals().get("vorname2",""), locals().get("nachname2",""), locals().get("ausweis2",""), locals().get("beginn2",""), locals().get("ende2",""), locals().get("vorhaltung2",""),
            locals().get("vorname3",""), locals().get("nachname3",""), locals().get("ausweis3",""), locals().get("beginn3",""), locals().get("ende3",""), locals().get("vorhaltung3",""),
            locals().get("vorname4",""), locals().get("nachname4",""), locals().get("ausweis4",""), locals().get("beginn4",""), locals().get("ende4",""), locals().get("vorhaltung4",""),
            locals().get("vorname5",""), locals().get("nachname5",""), locals().get("ausweis5",""), locals().get("beginn5",""), locals().get("ende5",""), locals().get("vorhaltung5",""),
            excel_name, json.dumps(payload, ensure_ascii=False)
        ))

    # DB feltöltés/frissítés Drive-ra
    sync_db_to_drive()

    headers = {
        "Content-Disposition": f'attachment; filename="{excel_name}"',
        "Content-Length": str(len(excel_bytes)),
        "Cache-Control": "no-store",
    }
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# ---------- PDF előnézet ----------
@app.post("/generate_pdf")
async def generate_pdf(
    request: Request,
    datum: str = Form(...),
    bau: str = Form(...),
    basf_beauftragter: str = Form(""),
    geraet: str = Form(""),
    beschrijving: str = Form(""),
    break_minutes: int = Form(60),
    vorname1: str = Form(""), nachname1: str = Form(""), ausweis1: str = Form(""), beginn1: str = Form(""), ende1: str = Form(""), vorhaltung1: str = Form(""),
    vorname2: str = Form(""), nachname2: str = Form(""), ausweis2: str = Form(""), beginn2: str = Form(""), ende2: str = Form(""), vorhaltung2: str = Form(""),
    vorname3: str = Form(""), nachname3: str = Form(""), ausweis3: str = Form(""), beginn3: str = Form(""), ende3: str = Form(""), vorhaltung3: str = Form(""),
    vorname4: str = Form(""), nachname4: str = Form(""), ausweis4: str = Form(""), beginn4: str = Form(""), ende4: str = Form(""), vorhaltung4: str = Form(""),
    vorname5: str = Form(""), nachname5: str = Form(""), beginn5: str = Form(""), ende5: str = Form(""), ausweis5: str = Form(""), vorhaltung5: str = Form(""),
):
    if not _is_user(request):
        return RedirectResponse("/login?next=/", status_code=303)
    if not (REPORTLAB_AVAILABLE and PIL_AVAILABLE):
        return PlainTextResponse("PDF előállítás nem elérhető (telepítsd: reportlab, pillow).", status_code=501)

    wb = load_workbook(os.path.join(os.getcwd(), "GP-t.xlsx")); ws = wb.active
    date_text = datum
    try:
        dt = datetime.strptime(datum.strip(), "%Y-%m-%d"); date_text = dt.strftime("%d.%m.%Y")
    except Exception:
        pass

    r1, c1, r2, c2 = find_description_block(ws)
    workers = []
    for i in range(1, 6):
        vn = locals().get(f"vorname{i}", "") or ""
        nn = locals().get(f"nachname{i}", "") or ""
        aw = locals().get(f"ausweis{i}", "") or ""
        bg = locals().get(f"beginn{i}", "") or ""
        en = locals().get(f"ende{i}", "") or ""
        vh = locals().get(f"vorhaltung{i}", "") or ""
        if not (vn or nn or aw or bg or en or vh): continue
        workers.append((vn, nn, aw, bg, en, vh))

    total_hours = 0.0
    for (_, _, _, bg, en, _) in workers:
        hb = parse_hhmm(bg); he = parse_hhmm(en); total_hours += hours_with_breaks(hb, he, int(break_minutes))

    pdf_bytes = _build_pdf_preview(
        date_text=date_text, bau=bau, basf_beauftragter=basf_beauftragter, beschreibung=beschrijving,
        ws=ws, r1=r1, c1=c1, r2=r2, c2=c2, workers=workers, total_hours=total_hours
    )

    fname = f"leistungsnachweis_preview_{uuid.uuid4().hex[:8]}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"', "Content-Length": str(len(pdf_bytes)), "Cache-Control": "no-store"}
    return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)

# ---------- TRANSLATE: AZURE + (opcionális) LT fallback ----------
AZURE_EP  = os.getenv("AZURE_TRANSLATOR_ENDPOINT", "").strip().rstrip("/")
AZURE_KEY = os.getenv("AZURE_TRANSLATOR_KEY", "").strip()
AZURE_RG  = os.getenv("AZURE_TRANSLATOR_REGION", "").strip()

LT_ENDPOINT = os.getenv("LT_ENDPOINT", "").strip()
LT_BACKUP_ENDPOINT = os.getenv("LT_BACKUP_ENDPOINT", "").strip()
LT_VIRTUAL_HOST = os.getenv("LT_VIRTUAL_HOST", "libretranslate.com").strip()
try:
    LT_TIMEOUT = float(os.getenv("LT_TIMEOUT", "12"))
except Exception:
    LT_TIMEOUT = 12.0

_IP_RE = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
def _headers_for(endpoint: str, base_headers: dict | None = None) -> dict:
    base = dict(base_headers or {})
    host = (urlparse(endpoint).hostname or "").strip()
    if _IP_RE.match(host):
        base.setdefault("Host", LT_VIRTUAL_HOST or "libretranslate.com")
    return base

def azure_ready() -> bool:
    return bool(AZURE_EP and AZURE_KEY and AZURE_RG)

def _azure_translate(text: str, source: str, target: str, timeout: float = 12.0) -> str:
    if not azure_ready():
        raise RuntimeError("Azure translator not configured")
    url = f"{AZURE_EP}/translate"
    qs = {"api-version": "3.0", "from": source, "to": target}
    headers = {
        "Ocp-Apim-Subscription-Key": AZURE_KEY,
        "Ocp-Apim-Subscription-Region": AZURE_RG,
        "Content-Type": "application/json; charset=UTF-8",
        "Accept": "application/json",
        "User-Agent": "pdf-edit/azure"
    }
    body = [{"text": text}]
    with httpx.Client(timeout=timeout, follow_redirects=True) as cli:
        r = cli.post(f"{url}?{urlencode(qs)}", headers=headers, json=body)
    if r.status_code != 200:
        raise RuntimeError(f"Azure HTTP {r.status_code}: {(r.text or '')[:200]}")
    jr = r.json()
    try:
        return jr[0]["translations"][0]["text"]
    except Exception:
        raise RuntimeError(f"Azure parse error: {jr}")

def _lt_translate(text: str, source: str, target: str, timeout: float = 12.0) -> str:
    eps = [e for e in [LT_ENDPOINT, LT_BACKUP_ENDPOINT] if e]
    if not eps:
        raise RuntimeError("LibreTranslate endpoint not configured")
    base_headers = {"Accept": "application/json", "Content-Type": "application/json", "User-Agent": "pdf-edit/lt"}
    data = {"q": text, "source": source, "target": target, "format": "text"}
    last = None
    for ep in eps:
        try:
            headers = _headers_for(ep, base_headers)
            with httpx.Client(timeout=timeout, follow_redirects=True) as cli:
                r = cli.post(ep, headers=headers, json=data)
            if r.status_code == 200:
                jr = r.json()
                for key in ("translatedText", "translation", "translated"):
                    if isinstance(jr, dict) and isinstance(jr.get(key), str) and jr.get(key):
                        return jr[key]
                raise RuntimeError(f"LT parse error: {jr}")
            last = f"LT HTTP {r.status_code}: {(r.text or '')[:200]}"
        except Exception as e:
            last = repr(e)
    raise RuntimeError(last or "LT unknown error")

@app.post("/api/translate")
async def api_translate(payload: dict = Body(...)):
    text   = (payload.get("text") or "").strip()
    source = (payload.get("source") or "hr").strip() or "hr"
    target = (payload.get("target") or "de").strip() or "de"
    if not text:
        return JSONResponse({"translated": ""})

    # 1) Azure first
    try:
        tr = _azure_translate(text, source, target, timeout=LT_TIMEOUT)
        return JSONResponse({"translated": tr, "engine": "azure"})
    except Exception as e1:
        err_az = repr(e1)

    # 2) Optional LT fallback
    try:
        tr = _lt_translate(text, source, target, timeout=LT_TIMEOUT)
        return JSONResponse({"translated": tr, "engine": "libretranslate", "azure_error": err_az})
    except Exception as e2:
        err_lt = repr(e2)

    return JSONResponse({"error": f"Translator error. azure={err_az}; lt={err_lt}"}, status_code=502)

# ---------- DIAG ----------
@app.get("/api/translator_info")
async def translator_info():
    out = {
        "azure": {
            "configured": azure_ready(),
            "endpoint": AZURE_EP,
            "region": AZURE_RG[:8] + "..." if AZURE_RG else "",
            "key_present": bool(AZURE_KEY)
        },
        "lt": {
            "primary": LT_ENDPOINT, "backup": LT_BACKUP_ENDPOINT, "virtual_host": LT_VIRTUAL_HOST, "timeout": LT_TIMEOUT
        }
    }
    dns = []
    for ep in [LT_ENDPOINT, LT_BACKUP_ENDPOINT]:
        if not ep: continue
        host = ""
        try:
            host = ep.split("://",1)[1].split("/",1)[0]
        except Exception: pass
        if not host:
            dns.append({"endpoint": ep, "host": "", "dns_ok": False, "ips": [], "error": "bad url"}); continue
        try:
            infos = socket.getaddrinfo(host, 443, proto=socket.IPPROTO_TCP)
            ips = sorted({ai[4][0] for ai in infos})
            dns.append({"endpoint": ep, "host": host, "dns_ok": True, "ips": ips})
        except Exception as e:
            dns.append({"endpoint": ep, "host": host, "dns_ok": False, "ips": [], "error": repr(e)})
    out["lt_dns"] = dns
    return JSONResponse(out)

@app.post("/api/translate_probe")
async def translate_probe(payload: dict = Body(...)):
    text   = (payload.get("text") or "").strip()
    source = (payload.get("source") or "hr").strip() or "hr"
    target = (payload.get("target") or "de").strip() or "de"
    if not text:
        return JSONResponse({"ok": True, "note": "empty text"})

    trace = []
    a = {"engine": "azure", "endpoint": AZURE_EP, "region": AZURE_RG, "configured": azure_ready()}
    try:
        if azure_ready():
            tr = _azure_translate(text, source, target, timeout=LT_TIMEOUT)
            a["ok"] = True; a["translated"] = tr; trace.append(a)
            return JSONResponse({"ok": True, "engine_used": "azure", "translated": tr, "trace": trace})
        else:
            a["ok"] = False; a["error"] = "not configured"; trace.append(a)
    except Exception as e:
        a["ok"] = False; a["error"] = repr(e); trace.append(a)

    l = {"engine": "libretranslate", "endpoints": [LT_ENDPOINT, LT_BACKUP_ENDPOINT]}
    try:
        tr = _lt_translate(text, source, target, timeout=LT_TIMEOUT)
        l["ok"] = True; l["translated"] = tr; trace.append(l)
        return JSONResponse({"ok": True, "engine_used": "libretranslate", "translated": tr, "trace": trace})
    except Exception as e:
        l["ok"] = False; l["error"] = repr(e); trace.append(l)

    return JSONResponse({"ok": False, "trace": trace}, status_code=502)

# ===================== ADMIN =====================

@app.get("/admin", response_class=HTMLResponse)
async def admin_index(request: Request, q_bau: str = "", q_date: str = ""):
    if not _is_admin(request):
        return RedirectResponse("/admin/login?next=/admin", status_code=303)
    sql = "SELECT id, created_at, datum, bau, basf_beauftragter, excel_filename FROM submissions "
    clauses = []; params = []
    if q_bau.strip():
        clauses.append("bau LIKE ?"); params.append(f"%{q_bau.strip()}%")
    if q_date.strip():
        clauses.append("datum LIKE ?"); params.append(f"%{q_date.strip()}%")
    if clauses: sql += "WHERE " + " AND ".join(clauses) + " "
    sql += "ORDER BY created_at DESC LIMIT 200"
    with db_conn() as c:
        rows = c.execute(sql, params).fetchall()
    return templates.TemplateResponse("admin.html", {"request": request, "rows": rows, "q_bau": q_bau, "q_date": q_date})

@app.get("/admin/view/{sid}", response_class=HTMLResponse)
async def admin_view(request: Request, sid: int):
    if not _is_admin(request):
        return RedirectResponse(f"/admin/login?next=/admin/view/{sid}", status_code=303)
    with db_conn() as c:
        row = c.execute("SELECT * FROM submissions WHERE id = ?", (sid,)).fetchone()
        if not row:
            return PlainTextResponse("Nincs ilyen bejegyzés.", status_code=404)
    try:
        payload = json.loads(row["payload_json"] or "{}")
    except Exception:
        payload = {}
    return templates.TemplateResponse("admin_detail.html", {"request": request, "sub": row, "payload": payload})

@app.get("/download/{fname}")
async def download_file(request: Request, fname: str):
    if not _is_admin(request):
        return RedirectResponse(f"/admin/login?next=/download/{fname}", status_code=303)
    fp = GEN_DIR / fname
    if not fp.exists():
        return PlainTextResponse("Fájl nem található.", status_code=404)
    data = fp.read_bytes()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )

# ---------- Health ----------
@app.get("/healthz")
async def healthz():
    return {"ok": True}

# ---------- Main ----------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT", "10000")), reload=True)
    
